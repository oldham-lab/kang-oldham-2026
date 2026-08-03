# Does the exact same thing as the R script but in python

import os
import pyreadr
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# Note: .qs files cannot be read directly in Python.
# Re-save your R objects as .rds first using:
#   qs::qsave -> saveRDS, or add this to your R pipeline:
#   saveRDS(result_list, gsub(".qs", ".rds", save_path))
# Then point the paths below to the .rds files.
# ============================================================

# ============================================================
# INPUT / OUTPUT PATHS  —  edit these before running
# ============================================================

EDGER_PATHS = {
    ("jorstad", "DFC"): os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/edgeR/jorstad/edgeR_jorstad_DFC_subclass_matchedGenes.rds"),
    ("jorstad", "MTG"): os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/edgeR/jorstad/edgeR_jorstad_MTG_subclass_matchedGenes.rds"),
    ("seaad",   "DFC"): os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/edgeR/seaad/edgeR_seaad_DFC_subclass_matchedGenes.rds"),
    ("seaad",   "MTG"): os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/edgeR/seaad/edgeR_seaad_MTG_subclass_matchedGenes.rds"),
}

DESEQ2_PATHS = {
    ("jorstad", "DFC"): os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/DESeq2/jorstad/DESeq2_jorstad_DFC_subclass_matchedGenes.rds"),
    ("jorstad", "MTG"): os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/DESeq2/jorstad/DESeq2_jorstad_MTG_subclass_matchedGenes.rds"),
    ("seaad",   "DFC"): os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/DESeq2/seaad/DESeq2_seaad_DFC_subclass_matchedGenes.rds"),
    ("seaad",   "MTG"): os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/DESeq2/seaad/DESeq2_seaad_MTG_subclass_matchedGenes.rds"),
}

OUT_DIR     = os.path.join(os.environ.get("EXTFIG_DIR", "/home/gugene/figures"), "figure_1/DE/tables")
PADJ_THRESH = 0.05
SIG_FIGS    = 4


# ============================================================
# Column legend definitions
# ============================================================
EDGER_LEGEND = [
    ("celltype",  "Cell subclass for which DE was tested (one-vs-all contrast)"),
    ("genes",     "Gene symbol"),
    ("logFC",     "Log2 fold change: positive = higher in this celltype vs all others"),
    ("logCPM",    "Log2 counts per million (average expression across all samples)"),
    ("LR",        "Likelihood ratio statistic from the generalized linear model test"),
    ("PValue",    "Nominal p-value from the likelihood ratio test"),
    ("FDR",       "False discovery rate-adjusted p-value (Benjamini-Hochberg); "
                  "rows are filtered to FDR < threshold"),
]

DESEQ2_LEGEND = [
    ("celltype",        "Cell subclass for which DE was tested (one-vs-all contrast)"),
    ("baseMean",        "Mean normalized count across all samples"),
    ("log2FoldChange",  "Log2 fold change: positive = higher in this celltype vs all others"),
    ("lfcSE",           "Standard error of the log2 fold change estimate"),
    ("stat",            "Wald or LRT statistic"),
    ("pvalue",          "Nominal p-value"),
    ("padj",            "Adjusted p-value (Benjamini-Hochberg); "
                        "rows are filtered to padj < threshold"),
]


# ============================================================
# Styles
# ============================================================
FONT_NAME = "Arial"
FONT_SIZE = 11

def header_font():
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")

def section_font():
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")

def body_font():
    return Font(name=FONT_NAME, size=FONT_SIZE)

def italic_font():
    return Font(name=FONT_NAME, size=10, italic=True)

def bold_font():
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True)

def header_fill():
    return PatternFill("solid", fgColor="D9E1F2")

def section_fill():
    return PatternFill("solid", fgColor="4472C4")

def center_align():
    return Alignment(horizontal="center", vertical="center")

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def bottom_border():
    return Border(bottom=Side(style="medium", color="4472C4"))


# ============================================================
# Helper functions
# ============================================================

def load_rds_list(path):
    """Load an RDS file containing a named list of data frames."""
    result = pyreadr.read_r(path)
    # pyreadr returns an OrderedDict; the named list becomes individual keys
    return result


def compile_de_results(reslist, padj_col):
    """Filter each celltype df to significant genes, prepend celltype column, concatenate."""
    frames = []
    for ct, df in reslist.items():
        df = df[df[padj_col].notna() & (df[padj_col] < PADJ_THRESH)].copy()
        df.insert(0, "celltype", ct)
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def round_sig(x, n=SIG_FIGS):
    """Round a float to n significant figures."""
    if pd.isna(x) or not isinstance(x, float):
        return x
    from math import log10, floor
    if x == 0:
        return 0
    d = n - int(floor(log10(abs(x)))) - 1
    return round(x, d)


def round_sig_df(df, n=SIG_FIGS):
    """Apply significant-figure rounding to all float columns."""
    df = df.copy()
    for col in df.select_dtypes(include="float").columns:
        df[col] = df[col].apply(lambda x: round_sig(x, n))
    return df


def write_df_to_sheet(ws, df):
    """Write a DataFrame to a worksheet with header + body formatting."""
    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        cell.fill      = header_fill()
        cell.alignment = center_align()
        cell.border    = bottom_border()

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font()
            cell.alignment = center_align()

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_values  = [str(col_name)] + [str(v) for v in df.iloc[:, col_idx - 1]]
        max_width   = min(max(len(v) for v in col_values) + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_width

    # Freeze top row
    ws.freeze_panes = "A2"


def write_legend_section(ws, legend_rows, start_row, section_title):
    """Write a labelled legend section (title row + column/description pairs)."""
    # Section header spanning both columns
    title_cell = ws.cell(row=start_row, column=1, value=section_title)
    title_cell.font      = section_font()
    title_cell.fill      = section_fill()
    title_cell.alignment = left_align()
    ws.cell(row=start_row, column=2).fill      = section_fill()
    ws.cell(row=start_row, column=2).alignment = left_align()

    # Column header row
    for col_idx, label in enumerate(["Column", "Description"], start=1):
        cell = ws.cell(row=start_row + 1, column=col_idx, value=label)
        cell.font      = bold_font()
        cell.fill      = header_fill()
        cell.alignment = left_align()
        cell.border    = bottom_border()

    # Legend entries
    for i, (col_name, description) in enumerate(legend_rows):
        name_cell = ws.cell(row=start_row + 2 + i, column=1, value=col_name)
        name_cell.font      = bold_font()
        name_cell.alignment = left_align(wrap=True)

        desc_cell = ws.cell(row=start_row + 2 + i, column=2, value=description)
        desc_cell.font      = body_font()
        desc_cell.alignment = left_align(wrap=True)

    return start_row + 2 + len(legend_rows)  # return next available row


def build_workbook(edger_path, deseq2_path, label):
    print(f"Building: {label}")

    edger_list  = load_rds_list(edger_path)
    deseq2_list = load_rds_list(deseq2_path)

    edger_df  = round_sig_df(compile_de_results(edger_list,  padj_col="FDR"))
    deseq2_df = round_sig_df(compile_de_results(deseq2_list, padj_col="padj"))

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # --- edgeR tab ---
    ws_edger = wb.create_sheet("edgeR")
    write_df_to_sheet(ws_edger, edger_df)

    # --- DESeq2 tab ---
    ws_deseq2 = wb.create_sheet("DESeq2")
    write_df_to_sheet(ws_deseq2, deseq2_df)

    # --- Legend tab ---
    ws_legend = wb.create_sheet("Legend")
    ws_legend.column_dimensions["A"].width = 20
    ws_legend.column_dimensions["B"].width = 70
    ws_legend.freeze_panes = "A2"

    next_row = write_legend_section(ws_legend, EDGER_LEGEND,  start_row=1,
                                    section_title="edgeR columns")
    next_row = write_legend_section(ws_legend, DESEQ2_LEGEND, start_row=next_row + 2,
                                    section_title="DESeq2 columns")

    # Filtering note
    note_row  = next_row + 2
    note_cell = ws_legend.cell(row=note_row, column=1,
                                value=f"Note: All results are filtered to adjusted p-value < {PADJ_THRESH}.")
    note_cell.font      = italic_font()
    note_cell.alignment = left_align()

    out_path = os.path.join(OUT_DIR, f"DE_results_{label}.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


# ============================================================
# Build all four workbooks
# ============================================================
os.makedirs(OUT_DIR, exist_ok=True)

for (dataset, region) in [("jorstad", "DFC"), ("jorstad", "MTG"),
                           ("seaad",   "DFC"), ("seaad",   "MTG")]:
    build_workbook(
        edger_path  = EDGER_PATHS[(dataset, region)],
        deseq2_path = DESEQ2_PATHS[(dataset, region)],
        label       = f"{dataset}_{region}"
    )

print("All workbooks complete.")
